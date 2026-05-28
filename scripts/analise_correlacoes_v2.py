"""V2: analise refinada com skiprows correto e cruzamento adicional (cargo).

Foca nos achados:
- SICA_RA e SIGOT precisam de skiprows (4 e 2)
- Validar CPFs mascarados no SICA_RA
- Cruzar CARGO entre RH e matrizes de perfil esperado
- Cruzar CC entre RH e Mapeamento CCO_CSC
"""
import csv
import re
from pathlib import Path

import openpyxl

ORIGEM = Path("Arquivos_origem")


def norm_cpf(v):
    if v is None:
        return ""
    s = re.sub(r"\D", "", str(v))
    return s.zfill(11) if s else ""


def cpf_mascarado(v):
    if v is None:
        return False
    s = str(v).strip().upper()
    return "X" in s


def ler_csv_skip(arq, encoding="latin-1", sep=";", skip=0):
    with open(arq, "r", encoding=encoding, newline="") as f:
        rows = list(csv.reader(f, delimiter=sep))
    return rows[skip:]


def ler_xlsx(arq, aba_idx=0):
    wb = openpyxl.load_workbook(arq, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[aba_idx]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


# ============================================================
# RH
# ============================================================
print("="*70)
print("BASE RH")
print("="*70)

rh_ativos = ler_csv_skip(ORIGEM / "PROJETOIAM (8).CSV")
header_rh = rh_ativos[0]
print(f"RH ATIVOS: {len(rh_ativos)-1} linhas")
# Identifica coluna por nome
def idx(header, nome):
    for i, c in enumerate(header):
        if c and c.strip() == nome:
            return i
    return -1

i_cpf = idx(header_rh, "Numero do CPF")
i_mat = idx(header_rh, "Matricula")
i_cargo_cod = idx(header_rh, "C\xf3digo do Cargo")
i_cargo_desc = idx(header_rh, "Descritivo do Cargo")
i_cc_cod = idx(header_rh, "C\xf3digo do Centro de Custo")
i_cc_nome = idx(header_rh, "Nome do Centro de Custo")
print(f"  Colunas: cpf={i_cpf}, mat={i_mat}, cargo_cod={i_cargo_cod}, cargo_desc={i_cargo_desc}, cc={i_cc_cod}, cc_nome={i_cc_nome}")

cpfs_rh_at = set()
mat_rh_at = set()
cargos_rh_at = set()
ccs_rh_at = set()
cpf_to_cargo = {}
cpf_to_cc = {}
cpf_to_mat = {}

for r in rh_ativos[1:]:
    if len(r) <= i_cargo_desc:
        continue
    cpf = norm_cpf(r[i_cpf])
    if cpf:
        cpfs_rh_at.add(cpf)
        mat_rh_at.add(r[i_mat].strip() if r[i_mat] else "")
        cargo = (r[i_cargo_desc] or "").strip().upper()
        cc = (r[i_cc_nome] or "").strip()
        if cargo:
            cargos_rh_at.add(cargo)
            cpf_to_cargo[cpf] = cargo
        if cc:
            ccs_rh_at.add(cc)
            cpf_to_cc[cpf] = cc
        cpf_to_mat[cpf] = r[i_mat].strip() if r[i_mat] else ""

print(f"  CPFs distintos: {len(cpfs_rh_at)}")
print(f"  Cargos distintos: {len(cargos_rh_at)}")
print(f"  CCs distintos: {len(ccs_rh_at)}")
print(f"  Ex. cargos: {list(cargos_rh_at)[:3]}")
print(f"  Ex. CCs: {list(ccs_rh_at)[:3]}")

# Desligados
rh_desl = ler_csv_skip(ORIGEM / "PROJETOIAMDESLIGADOS (2).CSV")
h_desl = rh_desl[0]
i_cpf_d = idx(h_desl, "Numero do CPF")
i_data_d = idx(h_desl, "Data do Desligamento")
cpfs_rh_desl = set()
for r in rh_desl[1:]:
    cpf = norm_cpf(r[i_cpf_d]) if len(r) > i_cpf_d else ""
    if cpf:
        cpfs_rh_desl.add(cpf)
print(f"\nRH DESLIGADOS: {len(rh_desl)-1} linhas, {len(cpfs_rh_desl)} CPFs distintos")

# ============================================================
# SICA_RA (skiprows=4) - CPF pode estar mascarado!
# ============================================================
print("\n" + "="*70)
print("SICA_RA (com skiprows=4)")
print("="*70)
sica_ra = ler_csv_skip(ORIGEM / "SICA_RA_30_04.csv", skip=4)
print(f"Total linhas (apos skip): {len(sica_ra)}")
print(f"Cabecalho:")
for i, c in enumerate(sica_ra[0]):
    print(f"  [{i:2d}] {c!r}")

i_cpf_sr = idx(sica_ra[0], "CPF")
print(f"\nColuna CPF no indice: {i_cpf_sr}")

cpfs_sr_mascarados = 0
cpfs_sr_normais = 0
cpfs_sr_set = set()
for r in sica_ra[1:]:
    if len(r) <= i_cpf_sr:
        continue
    raw = r[i_cpf_sr]
    if cpf_mascarado(raw):
        cpfs_sr_mascarados += 1
    else:
        cpfs_sr_normais += 1
        n = norm_cpf(raw)
        if n and len(n) == 11:
            cpfs_sr_set.add(n)

print(f"  Linhas com CPF MASCARADO (XXX): {cpfs_sr_mascarados}")
print(f"  Linhas com CPF normal: {cpfs_sr_normais}")
print(f"  CPFs unicos validos: {len(cpfs_sr_set)}")
if cpfs_sr_set:
    inter = cpfs_sr_set & cpfs_rh_at
    print(f"  Bate com RH ATIVOS: {len(inter)} ({100*len(inter)/len(cpfs_sr_set):.1f}%)")

# ============================================================
# SIGOT (skiprows=2)
# ============================================================
print("\n" + "="*70)
print("SIGOT (com skiprows=2)")
print("="*70)
sigot = ler_csv_skip(ORIGEM / "SIGOT_30_04.csv", skip=2)
print(f"Total linhas (apos skip): {len(sigot)}")
print(f"Cabecalho:")
for i, c in enumerate(sigot[0][:20]):
    print(f"  [{i:2d}] {c!r}")

i_cpf_sg = idx(sigot[0], "CPF")
print(f"\nColuna CPF no indice: {i_cpf_sg}")

cpfs_sg_mascarados = 0
cpfs_sg_set = set()
for r in sigot[1:]:
    if len(r) <= i_cpf_sg:
        continue
    raw = r[i_cpf_sg]
    if cpf_mascarado(raw):
        cpfs_sg_mascarados += 1
    else:
        n = norm_cpf(raw)
        if n and len(n) == 11:
            cpfs_sg_set.add(n)

print(f"  Linhas com CPF MASCARADO: {cpfs_sg_mascarados}")
print(f"  CPFs unicos validos: {len(cpfs_sg_set)}")
if cpfs_sg_set:
    inter = cpfs_sg_set & cpfs_rh_at
    print(f"  Bate com RH ATIVOS: {len(inter)} ({100*len(inter)/len(cpfs_sg_set):.1f}%)")

# ============================================================
# CARGO: RH x Matrizes de perfil esperado
# ============================================================
print("\n" + "="*70)
print("CRUZAMENTO CARGO: RH ATIVOS vs Matrizes de Perfil Esperado")
print("="*70)

def cargos_matriz(arq):
    rows = ler_xlsx(arq)
    h = [str(c).strip() if c else "" for c in rows[0]]
    i = idx(h, "CARGO")
    if i < 0:
        return None, set()
    cargos = set()
    for r in rows[1:]:
        if len(r) > i and r[i]:
            cargos.add(str(r[i]).strip().upper())
    return i, cargos

matrizes = [
    "MATRIZ DE PERFIL DE ACESSO - SIGOT.xlsx",
    "MATRIZ DE PERFIL DE ACESSO SICA ESFERA.xlsx",
    "MATRIZ DE PERFIL DE ACESSO SICA RA.xlsx",
    "MATRIZ DE PERFIL DE ACESSO SYSTUR.xlsx",
    "Matriz de Perfil de Acessso - IC Integrador Contabil.xlsx",
]
for m in matrizes:
    _, cargos = cargos_matriz(ORIGEM / m)
    inter = cargos & cargos_rh_at
    print(f"\n{m}:")
    print(f"  Cargos na matriz: {len(cargos)}")
    print(f"  Cargos no RH:     {len(cargos_rh_at)}")
    print(f"  INTERSECAO (cargos que a matriz reconhece no RH): {len(inter)}")
    print(f"  Cobertura da matriz sobre RH: {100*len(inter)/len(cargos_rh_at):.1f}%")
    fora = cargos - cargos_rh_at
    if fora:
        print(f"  Cargos NA MATRIZ que NAO existem no RH: {len(fora)}")
        print(f"    Exemplos: {list(fora)[:3]}")

# ============================================================
# CCO/CSC: RH x Mapeamento CCO_CSC
# ============================================================
print("\n" + "="*70)
print("CRUZAMENTO CC: RH ATIVOS vs Mapeamento CCO_CSC")
print("="*70)

cco = ler_xlsx(ORIGEM / "Mapeamento CCO_CSC (1).xlsx")
h = [str(c).strip() if c else "" for c in cco[0]]
i_cc = idx(h, "C\xf3digo do Centro de Custo")
i_func = idx(h, "Fun\xe7\xe3o")
i_sist = idx(h, "Sistemas")
i_perf = idx(h, "Perfis")
print(f"Colunas CCO: cc={i_cc}, funcao={i_func}, sistemas={i_sist}, perfis={i_perf}")

ccs_cco = set()
sistemas_cco = set()
for r in cco[1:]:
    if len(r) > i_cc and r[i_cc]:
        ccs_cco.add(str(r[i_cc]).strip())
    if len(r) > i_sist and r[i_sist]:
        for s in str(r[i_sist]).split(","):
            sistemas_cco.add(s.strip().upper())

print(f"  CCs no Mapeamento: {len(ccs_cco)}")
print(f"  Ex CCs cco: {list(ccs_cco)[:3]}")
print(f"  Sistemas distintos no Mapeamento: {sorted(sistemas_cco)}")

# CC do RH (ativos) - eu usei cc_nome anteriormente errado, eh o codigo de fato
# Vou usar i_cc_cod
ccs_rh_codigo = set()
for r in rh_ativos[1:]:
    if len(r) > i_cc_cod and r[i_cc_cod]:
        ccs_rh_codigo.add(str(r[i_cc_cod]).strip())
print(f"  CCs no RH (ativos): {len(ccs_rh_codigo)}")
print(f"  Ex CCs RH: {list(ccs_rh_codigo)[:3]}")
inter_cc = ccs_cco & ccs_rh_codigo
print(f"  INTERSECAO: {len(inter_cc)}")
print(f"  Cobertura do Mapeamento sobre RH: {100*len(inter_cc)/len(ccs_rh_codigo):.1f}%")

# ============================================================
# SIG: cruzar CPF SIG com cargo do RH (pra saber quais cargos os usuarios SIG tem)
# ============================================================
print("\n" + "="*70)
print("SIG x RH: quais cargos os usuarios do SIG tem (via CPF)?")
print("="*70)
sig = ler_xlsx(ORIGEM / "SIG_18.05.26.xlsx")
i_cpf_sig = idx([str(c) for c in sig[0]], "CPF")
cpfs_sig = set()
for r in sig[1:]:
    if len(r) > i_cpf_sig:
        n = norm_cpf(r[i_cpf_sig])
        if n:
            cpfs_sig.add(n)

cargos_sig = []
for cpf in cpfs_sig:
    c = cpf_to_cargo.get(cpf)
    if c:
        cargos_sig.append(c)

from collections import Counter
print(f"  Total CPFs no SIG: {len(cpfs_sig)}")
print(f"  Tem cargo no RH:   {len(cargos_sig)}")
print(f"\n  Top 15 cargos do RH dos usuarios SIG:")
for cargo, n in Counter(cargos_sig).most_common(15):
    print(f"    {n:4d}  {cargo}")
