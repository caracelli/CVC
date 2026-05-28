"""Analise sistematica de correlacoes entre todos os arquivos de Arquivos_origem.

Objetivo: mapear, para cada arquivo, quais chaves de cruzamento existem
(CPF, login, matricula, cargo, CC, perfil) e medir a interseccao entre eles.
"""
from pathlib import Path
import re
import csv
import sys
from collections import Counter

import openpyxl

ORIGEM = Path("Arquivos_origem")


def norm_cpf(v):
    if v is None:
        return ""
    s = re.sub(r"\D", "", str(v))
    if not s:
        return ""
    return s.zfill(11)


def ler_xlsx_aba(arq, aba=None, max_rows=None):
    wb = openpyxl.load_workbook(arq, read_only=True, data_only=True)
    nome = aba or wb.sheetnames[0]
    ws = wb[nome]
    linhas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows and i >= max_rows:
            break
        linhas.append(row)
    wb.close()
    return nome, linhas


def ler_csv(arq, encoding=None, sep=None):
    # Tentar diferentes encodings
    encodings = [encoding] if encoding else ["utf-8-sig", "latin-1", "cp1252"]
    seps = [sep] if sep else [";", ","]
    for enc in encodings:
        for s in seps:
            try:
                with open(arq, "r", encoding=enc, newline="") as f:
                    leitor = csv.reader(f, delimiter=s)
                    linhas = list(leitor)
                if len(linhas) > 1 and len(linhas[0]) > 1:
                    return enc, s, linhas
            except Exception:
                continue
    return None, None, []


def sumario_arquivo(nome, header, linhas_dados, qt_data):
    print(f"\n{'='*70}")
    print(f"ARQUIVO: {nome}")
    print(f"  Linhas de dados: {qt_data}")
    print(f"  Total de colunas: {len(header)}")
    print(f"  Cabecalho:")
    for i, c in enumerate(header, 1):
        print(f"    [{i:3d}] {c!r}")


def extrair_col(linhas, header, padrao_regex, case_insensitive=True):
    """Retorna lista de valores de uma coluna identificada por regex no nome."""
    flags = re.IGNORECASE if case_insensitive else 0
    idx = None
    for i, c in enumerate(header):
        if c and re.search(padrao_regex, str(c), flags):
            idx = i
            break
    if idx is None:
        return None, []
    vals = [row[idx] for row in linhas if idx < len(row)]
    return header[idx], vals


def analisa_arquivo(arq):
    print(f"\n{'#'*70}")
    print(f"# {arq.name}")
    print('#'*70)
    try:
        if arq.suffix.lower() in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(arq, read_only=True, data_only=True)
            print(f"  Abas: {wb.sheetnames}")
            for aba in wb.sheetnames:
                ws = wb[aba]
                print(f"  -> Aba '{aba}': {ws.max_row} linhas x {ws.max_column} colunas")
            # le a primeira aba
            nome_aba, linhas = ler_xlsx_aba(arq)
            wb.close()
            if not linhas:
                print("  (vazio)")
                return None
            # Detecta header: linha com mais valores nao-nulos nas primeiras 10
            best_idx = 0
            best_count = 0
            for i, row in enumerate(linhas[:10]):
                cnt = sum(1 for v in row if v is not None and str(v).strip())
                if cnt > best_count:
                    best_count = cnt
                    best_idx = i
            header = [str(c).strip() if c is not None else "" for c in linhas[best_idx]]
            dados = linhas[best_idx+1:]
        else:
            enc, sep, linhas = ler_csv(arq)
            print(f"  Encoding: {enc}, Sep: {sep!r}")
            if not linhas:
                print("  (vazio ou erro)")
                return None
            header = [c.strip() for c in linhas[0]]
            dados = linhas[1:]

        # Filtra linhas totalmente vazias
        dados = [r for r in dados if any(v is not None and str(v).strip() for v in r)]
        sumario_arquivo(arq.name, header, dados, len(dados))

        # Extrai chaves potenciais
        chaves = {}
        nome_col, vals = extrair_col(dados, header, r"^cpf|cpf\s*/|cpf\b")
        if nome_col:
            cpfs_norm = {norm_cpf(v) for v in vals if norm_cpf(v)}
            chaves["cpf"] = (nome_col, cpfs_norm, vals)
            print(f"\n  CPF: coluna {nome_col!r} -> {len(cpfs_norm)} CPFs distintos (normalizados)")
            ex = list(cpfs_norm)[:3]
            print(f"     Exemplos normalizados: {ex}")

        for chave, regex in [
            ("matricula", r"matr[ií]cula|chapa|cd_func"),
            ("login", r"login|usuario|cd_login|cd_usuario|user"),
            ("nome", r"^nome|nm_pessoa|nm_user|nome do colab"),
            ("cargo_cod", r"c[oó]d.*cargo|cd_cargo|cargo.*c[oó]d"),
            ("cargo_desc", r"cargo|fun[cç][aã]o|nm_cargo"),
            ("cc_cod", r"c[oó]d.*centro|cd_cc|centro.*c[oó]d|^cc$|^cco$|cd_centro"),
            ("cc_nome", r"centro.*custo|nm_centro|nm_cc"),
            ("sistema", r"^sistema|nm_sistema"),
            ("perfil", r"perfil|grupo|role|nm_role|nm_grupo|cd_grupo"),
            ("situacao", r"situa[cç][aã]o|status|st_|sit\b"),
            ("email", r"email|e-mail|cd_email"),
            ("desligamento", r"deslig|sa[ií]da|demiss|data.*deslig"),
        ]:
            if chave in chaves:
                continue
            nome_col, vals = extrair_col(dados, header, regex)
            if nome_col:
                distintos = {str(v).strip() for v in vals if v is not None and str(v).strip()}
                chaves[chave] = (nome_col, distintos, vals)
                print(f"  {chave}: coluna {nome_col!r} -> {len(distintos)} valores distintos")
                ex = list(distintos)[:3]
                print(f"     Exemplos: {ex}")

        return {"arquivo": arq.name, "header": header, "qt_dados": len(dados), "chaves": chaves}
    except Exception as e:
        print(f"  ERRO: {e}")
        import traceback
        traceback.print_exc()
        return None


def main():
    arquivos = sorted([p for p in ORIGEM.iterdir() if p.is_file() and not p.name.startswith("~$")])
    print(f"Total de arquivos a analisar: {len(arquivos)}")
    resultados = {}
    for arq in arquivos:
        r = analisa_arquivo(arq)
        if r:
            resultados[arq.name] = r

    # CRUZAMENTOS
    print("\n\n" + "="*70)
    print("CRUZAMENTOS DE CPF (interseccao com a base de RH ATIVOS)")
    print("="*70)

    # Identifica base de RH ativos (PROJETOIAM (8).CSV) e desligados
    rh_ativos = resultados.get("PROJETOIAM (8).CSV") or resultados.get("PROJETOIAM (8) (1).CSV")
    rh_desligados = resultados.get("PROJETOIAMDESLIGADOS (2).CSV")

    if rh_ativos and "cpf" in rh_ativos["chaves"]:
        cpfs_ativos = rh_ativos["chaves"]["cpf"][1]
        print(f"\nRH ATIVOS: {len(cpfs_ativos)} CPFs")
    else:
        cpfs_ativos = set()
        print("  RH Ativos sem coluna CPF identificada!")

    if rh_desligados and "cpf" in rh_desligados["chaves"]:
        cpfs_desligados = rh_desligados["chaves"]["cpf"][1]
        print(f"RH DESLIGADOS: {len(cpfs_desligados)} CPFs")
    else:
        cpfs_desligados = set()

    print(f"RH UNIVERSO (ativos+desligados): {len(cpfs_ativos | cpfs_desligados)} CPFs\n")

    for nome, r in resultados.items():
        if nome in ("PROJETOIAM (8).CSV", "PROJETOIAM (8) (1).CSV", "PROJETOIAMDESLIGADOS (2).CSV"):
            continue
        if "cpf" not in r["chaves"]:
            continue
        cpfs = r["chaves"]["cpf"][1]
        inter_ativ = cpfs & cpfs_ativos
        inter_desl = cpfs & cpfs_desligados
        sem_match = cpfs - cpfs_ativos - cpfs_desligados
        print(f"{nome}:")
        print(f"   CPFs no arquivo: {len(cpfs)}")
        if cpfs_ativos:
            print(f"   batem com ATIVOS:    {len(inter_ativ):4d}  ({100*len(inter_ativ)/len(cpfs):5.1f}%)")
        if cpfs_desligados:
            print(f"   batem com DESLIG:    {len(inter_desl):4d}  ({100*len(inter_desl)/len(cpfs):5.1f}%)")
        print(f"   ORFAOS (sem RH):     {len(sem_match):4d}  ({100*len(sem_match)/len(cpfs):5.1f}%)")
        if sem_match:
            ex = list(sem_match)[:3]
            print(f"     Exemplos orfaos: {ex}")
        print()


if __name__ == "__main__":
    main()
