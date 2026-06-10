# -*- coding: utf-8 -*-
"""EscritorExcel: aba TODAS + uma aba por tipo (rotulos), colunas e conteudo."""
import glob
import os
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from infraestrutura.escritores_arquivos.escritor_excel import EscritorExcel, _COLUNAS
from dominio.entidades.divergencia import Divergencia
from dominio.objetos_valor.sistema import Sistema
from dominio.objetos_valor.tipo_divergencia import TipoDivergencia


def _div(tipo, usuario, perfil_enc="P1"):
    return Divergencia(id=f"{tipo.value}-{usuario}", tipo=tipo,
                       sistema=Sistema.IC_INTEGRADOR_CONTABIL, usuario=usuario,
                       nome_usuario=usuario.upper(), descricao="desc",
                       perfil_encontrado=perfil_enc)


class TestEscritorExcel(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls._tmp = tempfile.mkdtemp(prefix="cvc_xls_")
        divs = [
            _div(TipoDivergencia.ACESSO_SEM_VINCULO_RH, "u1"),
            _div(TipoDivergencia.ACESSO_SEM_VINCULO_RH, "u2"),
            _div(TipoDivergencia.PERFIL_INVALIDO, "u3"),
        ]
        cls.caminho = EscritorExcel().salvar_divergencias(divs, cls._tmp)
        cls.wb = openpyxl.load_workbook(cls.caminho, read_only=True)

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()

    def _linhas(self, aba):
        return list(self.wb[aba].iter_rows(values_only=True))

    def test_arquivo_gerado_com_nome_padrao(self):
        self.assertTrue(os.path.basename(self.caminho).startswith("DIVERGENCIAS_"))
        self.assertTrue(self.caminho.endswith(".xlsx"))

    def test_aba_todas_tem_todas_as_linhas(self):
        linhas = self._linhas("TODAS")
        self.assertEqual(list(linhas[0]), _COLUNAS)        # cabecalho
        self.assertEqual(len(linhas) - 1, 3)               # 3 divergencias

    def test_uma_aba_por_tipo_com_rotulo(self):
        nomes = set(self.wb.sheetnames)
        self.assertIn("TODAS", nomes)
        self.assertIn("Sem Vinculo RH", nomes)
        self.assertIn("Perfil Invalido", nomes)

    def test_aba_do_tipo_filtra_so_aquele_tipo(self):
        linhas = self._linhas("Sem Vinculo RH")
        i_tipo = list(linhas[0]).index("Tipo")
        tipos = {r[i_tipo] for r in linhas[1:]}
        self.assertEqual(tipos, {"ACESSO_SEM_VINCULO_RH"})
        self.assertEqual(len(linhas) - 1, 2)               # u1, u2

    def test_conteudo_de_uma_linha(self):
        linhas = self._linhas("Perfil Invalido")
        hdr = list(linhas[0])
        row = dict(zip(hdr, linhas[1]))
        self.assertEqual(row["Usuario"], "u3")
        self.assertEqual(row["Sistema"], "IC_INTEGRADOR_CONTABIL")
        self.assertEqual(row["Perfil Encontrado"], "P1")


if __name__ == "__main__":
    unittest.main(verbosity=2)
